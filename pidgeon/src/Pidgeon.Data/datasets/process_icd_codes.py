#!/usr/bin/env python3
"""
ICD Code Processing Script

This script processes the ICD-9 and ICD-10 Excel files and extracts
the most common/important diagnosis codes for realistic message generation.

Usage:
    python process_icd_codes.py

Requirements:
    - pip install openpyxl (for Excel file reading)
    - section111validicd10-jan2025_0.xlsx
    - section111validicd9-jan2025_0.xlsx

Output:
    - common_icd10_codes.csv
    - common_icd9_codes.csv (for legacy systems)
    - icd_codes_summary.json
"""

import sys
import json
import csv
from collections import Counter, defaultdict

try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    print("❌ Error: openpyxl not installed")
    print("   Please run: pip install openpyxl")
    sys.exit(1)

def process_icd10_codes(excel_file, output_csv, max_codes=2000):
    """
    Process ICD-10 Excel file and extract most useful codes

    Strategy:
    1. Load all codes from Excel
    2. Categorize by first character/chapter
    3. Select most common codes per category
    4. Prioritize codes commonly used in ambulatory care
    """

    print(f"📊 Processing ICD-10 codes from {excel_file}...")

    try:
        workbook = load_workbook(excel_file, read_only=True)
        worksheet = workbook.active

        # Read all rows (assuming columns: Code, Description)
        codes = []
        categories = defaultdict(list)

        # Skip header row, process data
        for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), 1):
            if row_num % 10000 == 0:
                print(f"   📋 Processed {row_num:,} codes...")

            if len(row) >= 2 and row[0] and row[1]:
                code = str(row[0]).strip()
                description = str(row[1]).strip()

                if code and description:
                    # Categorize by ICD-10 chapter (first character)
                    chapter = get_icd10_chapter(code)

                    code_info = {
                        'code': code,
                        'description': description,
                        'chapter': chapter,
                        'priority': get_code_priority(code, description)
                    }

                    codes.append(code_info)
                    categories[chapter].append(code_info)

        print(f"   ✅ Total codes loaded: {len(codes):,}")
        print(f"   📚 Categories found: {len(categories)}")

        # Select top codes per category to ensure balanced coverage
        selected_codes = []
        codes_per_category = max_codes // len(categories) if categories else max_codes

        for chapter, chapter_codes in categories.items():
            # Sort by priority (high priority = low number)
            chapter_codes.sort(key=lambda x: x['priority'])

            # Take top codes for this category
            top_codes = chapter_codes[:codes_per_category]
            selected_codes.extend(top_codes)

            print(f"   📋 {chapter}: {len(top_codes)} codes selected")

        # Write to CSV
        print(f"💾 Writing {len(selected_codes)} codes to {output_csv}...")

        with open(output_csv, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = ['code', 'description', 'chapter', 'priority']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()

            for code_info in selected_codes:
                writer.writerow(code_info)

        return selected_codes

    except Exception as e:
        print(f"❌ Error processing ICD-10 file: {e}")
        return []

def get_icd10_chapter(code):
    """Determine ICD-10 chapter based on code prefix"""
    if not code:
        return "Unknown"

    first_char = code[0].upper()

    # ICD-10 chapter mapping
    chapter_map = {
        'A': 'Infectious diseases (A00-B99)',
        'B': 'Infectious diseases (A00-B99)',
        'C': 'Neoplasms (C00-D49)',
        'D': 'Neoplasms (C00-D49)',
        'E': 'Endocrine/Metabolic (E00-E89)',
        'F': 'Mental/Behavioral (F01-F99)',
        'G': 'Nervous system (G00-G99)',
        'H': 'Eye/Ear (H00-H95)',
        'I': 'Circulatory (I00-I99)',
        'J': 'Respiratory (J00-J99)',
        'K': 'Digestive (K00-K95)',
        'L': 'Skin (L00-L99)',
        'M': 'Musculoskeletal (M00-M99)',
        'N': 'Genitourinary (N00-N99)',
        'O': 'Pregnancy (O00-O9A)',
        'P': 'Perinatal (P00-P96)',
        'Q': 'Congenital (Q00-Q99)',
        'R': 'Symptoms/Signs (R00-R99)',
        'S': 'Injury/Poisoning (S00-T88)',
        'T': 'Injury/Poisoning (S00-T88)',
        'U': 'Special purposes (U00-U85)',
        'V': 'External causes (V00-Y99)',
        'W': 'External causes (V00-Y99)',
        'X': 'External causes (V00-Y99)',
        'Y': 'External causes (V00-Y99)',
        'Z': 'Health status (Z00-Z99)'
    }

    return chapter_map.get(first_char, f"Unknown ({first_char})")

def get_code_priority(code, description):
    """
    Assign priority to codes based on clinical importance
    Lower number = higher priority
    """

    # Very high priority (common ambulatory diagnoses)
    high_priority_codes = {
        'I10', 'E11.9', 'E78.5', 'Z00.00', 'Z23', 'J06.9',
        'R50.9', 'R05', 'R51.9', 'M54.5', 'J20.9', 'N39.0'
    }

    # High priority patterns
    high_priority_patterns = [
        'diabetes', 'hypertension', 'depression', 'anxiety', 'chest pain',
        'back pain', 'headache', 'fever', 'cough', 'screening', 'immunization',
        'covid', 'pneumonia', 'asthma', 'copd'
    ]

    # Medium priority patterns
    medium_priority_patterns = [
        'chronic', 'acute', 'unspecified', 'without complications',
        'primary', 'essential'
    ]

    code_lower = code.lower()
    desc_lower = description.lower()

    # Check exact code matches
    if code in high_priority_codes:
        return 1

    # Check description patterns
    for pattern in high_priority_patterns:
        if pattern in desc_lower:
            return 10

    # Z codes (health status) - often used
    if code.startswith('Z'):
        return 20

    # R codes (symptoms) - very common
    if code.startswith('R'):
        return 25

    # Common chronic conditions
    for pattern in medium_priority_patterns:
        if pattern in desc_lower:
            return 30

    # Three-character codes (less specific)
    if len(code) <= 3:
        return 50

    # Default priority
    return 100

def create_summary_json(icd10_codes, icd9_codes, output_file):
    """Create summary JSON with statistics and samples"""

    summary = {
        'created_date': str(datetime.now()),
        'icd10': {
            'total_codes': len(icd10_codes),
            'chapters': {}
        },
        'icd9': {
            'total_codes': len(icd9_codes)
        },
        'samples': {
            'high_priority_icd10': [],
            'common_chapters': {}
        }
    }

    # ICD-10 chapter breakdown
    chapter_counts = Counter(code['chapter'] for code in icd10_codes)
    summary['icd10']['chapters'] = dict(chapter_counts)

    # High priority samples
    high_priority = [code for code in icd10_codes if code['priority'] <= 10]
    summary['samples']['high_priority_icd10'] = high_priority[:20]

    # Sample codes by chapter
    for chapter in list(chapter_counts.keys())[:5]:
        chapter_codes = [code for code in icd10_codes if code['chapter'] == chapter]
        summary['samples']['common_chapters'][chapter] = chapter_codes[:10]

    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(summary, f, indent=2, ensure_ascii=False)

    return summary

def main():
    """Main execution function"""
    print("🏥 ICD Code Processing")
    print("=" * 40)

    # Check required files
    icd10_file = 'section111validicd10-jan2025_0.xlsx'
    icd9_file = 'section111validicd9-jan2025_0.xlsx'

    if not os.path.exists(icd10_file):
        print(f"❌ Error: {icd10_file} not found")
        return

    # Process ICD-10 codes
    icd10_codes = process_icd10_codes(icd10_file, 'common_icd10_codes.csv', max_codes=2000)

    if not icd10_codes:
        print("❌ Failed to process ICD-10 codes")
        return

    # Process ICD-9 codes (smaller file, usually)
    icd9_codes = []
    if os.path.exists(icd9_file):
        print(f"\n📊 Processing ICD-9 codes from {icd9_file}...")
        # Similar processing for ICD-9 (simpler structure)
        try:
            workbook = load_workbook(icd9_file, read_only=True)
            worksheet = workbook.active

            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if len(row) >= 2 and row[0] and row[1]:
                    code = str(row[0]).strip()
                    description = str(row[1]).strip()
                    if code and description:
                        icd9_codes.append({
                            'code': code,
                            'description': description
                        })

            # Write ICD-9 CSV
            with open('common_icd9_codes.csv', 'w', newline='', encoding='utf-8') as csvfile:
                fieldnames = ['code', 'description']
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                for code in icd9_codes[:1000]:  # Top 1000 ICD-9 codes
                    writer.writerow(code)

            print(f"   ✅ ICD-9 codes processed: {len(icd9_codes):,}")

        except Exception as e:
            print(f"⚠️  Warning: Could not process ICD-9 file: {e}")

    # Create summary
    from datetime import datetime
    import os

    summary = create_summary_json(icd10_codes, icd9_codes, 'icd_codes_summary.json')

    # Final report
    print(f"\n🎉 Processing Complete!")
    print(f"   📊 ICD-10 codes: {len(icd10_codes):,}")
    print(f"   📊 ICD-9 codes: {len(icd9_codes):,}")
    print(f"   📁 Files created:")
    print(f"      - common_icd10_codes.csv ({len(icd10_codes)} codes)")
    if icd9_codes:
        print(f"      - common_icd9_codes.csv ({min(len(icd9_codes), 1000)} codes)")
    print(f"      - icd_codes_summary.json")

    print(f"\n📋 Chapter distribution (ICD-10):")
    for chapter, count in Counter(code['chapter'] for code in icd10_codes).most_common(10):
        print(f"   {count:3d} codes - {chapter}")

    print(f"\n🚀 Next steps:")
    print(f"   1. Run: python load_data_to_sqlite.py")
    print(f"   2. Test with: SELECT COUNT(*) FROM diagnoses;")

if __name__ == '__main__':
    main()